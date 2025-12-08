#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import sqlite3
import os
import json
from datetime import datetime

# Files with real curriculum data
excel_files = [
    ('data/excel/مقرر مادة التربية الإسلامية للجذع مشترك.xlsx', 'جذع مشترك'),
    ('data/excel/مقرر مادة التربية الإسلامية ثانية بكالوريا.xlsx', 'ثانية بكالوريا'),
    ('data/excel/مقرر مادة التربية الإسلامية أولى بكالوريا.xlsx', 'أولى بكالوريا'),
]

def extract_templates_from_excel():
    """Extract lesson templates from Excel files"""
    templates = []
    
    for file_path, level in excel_files:
        if not os.path.exists(file_path):
            print(f"⚠️  File not found: {file_path}")
            continue
        
        print(f"\n📄 Processing: {os.path.basename(file_path)}")
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Get headers
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value] = col_idx
            
            print(f"   Headers: {list(headers.keys())}")
            
            # Extract data rows
            for row_idx in range(2, ws.max_row + 1):
                week_cell = ws.cell(row_idx, headers.get('رقم الأسبوع', 1))
                title_cell = ws.cell(row_idx, headers.get('عنوان الدرس', 2))
                course_cell = ws.cell(row_idx, headers.get('اسم المقرر', 3))
                level_cell = ws.cell(row_idx, headers.get('المستوى', 4))
                
                week = week_cell.value
                title = title_cell.value
                course = course_cell.value
                level_val = level_cell.value
                
                if title and title.strip():
                    template = {
                        'title': str(title).strip(),
                        'courseName': str(course).strip() if course else 'التربية الإسلامية',
                        'level': str(level_val).strip() if level_val else level,
                        'weekNumber': int(week) if week else None,
                        'description': '',
                        'estimatedSessions': 1,
                        'stages': [],
                        'scheduledSections': []
                    }
                    templates.append(template)
                    print(f"   ✓ Week {week}: {title} ({level_val or level})")
        
        except Exception as e:
            print(f"   ❌ Error processing {file_path}: {e}")
    
    return templates

def update_database(templates):
    """Update LessonTemplates in database"""
    db_path = 'classroom_dev.db'
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    
    print(f"\n🔄 Updating database with {len(templates)} templates...")
    
    try:
        # Delete existing templates
        c.execute('DELETE FROM LessonTemplates')
        deleted = c.rowcount
        print(f"   Deleted old templates: {deleted}")
        
        # Insert new templates
        for i, tpl in enumerate(templates, 1):
            c.execute('''
                INSERT INTO LessonTemplates 
                (title, description, estimatedSessions, stages, courseName, level, weekNumber, scheduledSections, createdAt, updatedAt)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                tpl['title'],
                tpl['description'],
                tpl['estimatedSessions'],
                json.dumps(tpl['stages'], ensure_ascii=False),
                tpl['courseName'],
                tpl['level'],
                tpl['weekNumber'],
                json.dumps(tpl['scheduledSections'], ensure_ascii=False),
                datetime.now().isoformat(),
                datetime.now().isoformat()
            ))
        
        conn.commit()
        print(f"   ✓ Inserted {len(templates)} new templates")
        
        # Verify
        c.execute('SELECT COUNT(*) FROM LessonTemplates')
        count = c.fetchone()[0]
        print(f"\n✅ Database now has {count} templates")
        
        # Show summary
        c.execute('''
            SELECT courseName, level, COUNT(*) cnt 
            FROM LessonTemplates 
            GROUP BY courseName, level 
            ORDER BY courseName, level
        ''')
        print("\n📊 Distribution:")
        for course, level, cnt in c.fetchall():
            print(f"   {course} / {level}: {cnt} templates")
        
        return True
    
    except Exception as e:
        print(f"❌ Error updating database: {e}")
        conn.rollback()
        return False
    
    finally:
        conn.close()

if __name__ == '__main__':
    print("="*60)
    print("🚀 Update LessonTemplates with Real Data")
    print("="*60)
    
    # Extract templates
    templates = extract_templates_from_excel()
    
    if templates:
        # Update database
        if update_database(templates):
            print("\n✨ SUCCESS! Templates updated successfully!")
    else:
        print("\n❌ No templates extracted from Excel files")
